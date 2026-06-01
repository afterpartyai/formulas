#!/usr/bin/env python
# -*- coding: UTF-8 -*-
#
# Copyright 2016-2026 European Commission (JRC);
# Licensed under the EUPL (the 'Licence');
# You may not use this work except in compliance with the Licence.
# You may obtain a copy of the Licence at: http://ec.europa.eu/idabc/eupl

"""
It provides Excel model class.

Sub-Modules:

.. currentmodule:: formulas.excel

.. autosummary::
    :nosignatures:
    :toctree: excel/

    ~cycle
    ~xlreader
"""
import tqdm
import os
import logging
import functools
import numpy as np
import os.path as osp
import schedula as sh
from ..ranges import Ranges
from ..errors import InvalidRangeName
from ..cell import Cell, RangesAssembler, Ref, CellWrapper, InvRangesAssembler
from ..tokens.operand import XlError, _re_sheet_id, _re_build_id
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
from openpyxl.utils.exceptions import IllegalCharacterError

log = logging.getLogger(__name__)
BOOK = sh.Token('Book')
SHEETS = sh.Token('Sheets')
CIRCULAR = sh.Token('CIRCULAR')

# Profiling support (set env FORMULAS_PROFILING=1 to enable).
import time as _ptime
import collections as _pcoll
_PROF = os.environ.get('FORMULAS_PROFILING', '0') == '1'
_prof = _pcoll.defaultdict(lambda: {'t': 0.0, 'n': 0})


def _prof_reset():
    _prof.clear()


def _prof_report():
    if not _prof:
        return
    print(f"\n{'=' * 80}")
    print("  FORMULAS LIBRARY INTERNAL PROFILE")
    print(f"{'=' * 80}")
    items = sorted(_prof.items(), key=lambda x: -x[1]['t'])
    print(f"  {'Operation':<45s} {'Time':>10s} {'Count':>8s} {'Avg':>10s}")
    print(f"  {'-' * 45} {'-' * 10} {'-' * 8} {'-' * 10}")
    for name, d in items:
        avg = d['t'] / d['n'] if d['n'] else 0
        print(f"  {name:<45s} {d['t']:>9.3f}s {d['n']:>8d} {avg:>9.4f}s")
    total = sum(d['t'] for d in _prof.values())
    print(f"  {'-' * 45} {'-' * 10}")
    print(f"  {'TOTAL profiled':<45s} {total:>9.3f}s")
    print(f"{'=' * 80}\n")


# ── Cell deactivation by font-color marker ──
# When env FORMULAS_DEACTIVATE_RGB is set (e.g. "FF123456"), any formula cell
# whose font color matches that RGB is substituted with a constant at
# compile time. This skips parse/AST/dependency-graph entry — ~48 KB RSS
# per skipped formula on tested workbooks.
#
# Frozen-value source order (first hit wins):
#   1. env FORMULAS_DEACTIVATE_VALUE if set (e.g. "0")
#   2. cell.value if openpyxl read it as data_only (rare in this code path)
#   3. 0
_DEACTIVATE_RGB = (os.environ.get('FORMULAS_DEACTIVATE_RGB') or '').upper().strip()
_DEACTIVATE_DEFAULT = os.environ.get('FORMULAS_DEACTIVATE_VALUE')


def _safe_color_rgb(color):
    """Returns hex RGB string for a cell's font color, or None if theme/indexed
    or otherwise non-RGB. Tolerant of openpyxl raising on .rgb access for
    theme colors."""
    if color is None:
        return None
    try:
        ctype = getattr(color, 'type', None)
        if ctype != 'rgb':
            return None
        rgb = color.rgb
        if not isinstance(rgb, str):
            return None
        return rgb.upper()
    except (TypeError, AttributeError):
        return None


def _is_deactivated_cell(cell):
    if not _DEACTIVATE_RGB:
        return False
    font = getattr(cell, 'font', None)
    if font is None:
        return False
    return _safe_color_rgb(font.color) == _DEACTIVATE_RGB


def _frozen_value_for(cell):
    if _DEACTIVATE_DEFAULT is not None:
        try:
            return float(_DEACTIVATE_DEFAULT)
        except ValueError:
            return _DEACTIVATE_DEFAULT
    return 0


# Progress callback and abort flag.  Set from the caller before calling
# loads/finish/calculate.
_progress_callback = None
_abort_flag = None


class ExcelAbortError(Exception):
    pass


def _check_abort():
    if _abort_flag and _abort_flag():
        raise ExcelAbortError("Load aborted by user")


def _report_progress(phase, current, total, detail=''):
    if _progress_callback:
        _progress_callback(phase, current, total, detail)


class XlCircular(XlError):
    def __str__(self):
        return '0'


ERR_CIRCULAR = XlCircular('#CIRC!')


# AcquiOS fork helpers for the iterative circular-reference solver.
# See specs/FORMULAS_ITERATIVE_CIRCULAR_SOLVER.md §3.1.

def _to_scalar(x):
    # Coerce a cell-function output (Ranges / numpy Array / scalar) to float.
    if hasattr(x, 'value'):
        x = x.value
    if hasattr(x, 'tolist'):
        x = x.tolist()
    while isinstance(x, list) and len(x) == 1:
        x = x[0]
    try:
        return float(x)
    except (TypeError, ValueError):
        return 0.0


def _make_scalar_ranges(qualified_name, value):
    # Build a Ranges object holding a single scalar value at the given address.
    # qualified_name e.g. "'[wb]Sheet'!A2"
    return Ranges().push(qualified_name, np.asarray([[value]], dtype=object))


def _read_scalar_from_dsp(dsp, key):
    # Read the current value of a non-SCC cell from the dispatcher's default
    # values. Returns 0.0 if absent (Excel's default seed convention).
    try:
        entry = dsp.default_values.get(key)
        if entry is None:
            return 0.0
        return _to_scalar(entry.get('value', 0.0))
    except (AttributeError, KeyError):
        return 0.0


def _get_name(name, names):
    if name not in names:
        name = name.upper()
        for n in names:
            if n.upper() == name:
                return n
    return name


def _encode_path(path):
    return path.replace('\\', '/')


def _decode_path(path):
    return path.replace('/', osp.sep)


def _book2dict(book):
    res = {}
    for ws in book.worksheets:
        s = res[ws.title.upper()] = {}
        for k, cell in ws._cells.items():
            value = getattr(cell, 'value', None)
            if value is not None:
                s[cell.coordinate] = value
    return res


def _res2books(res):
    return {k.upper(): _book2dict(v[BOOK]) for k, v in res.items()}


def _file2books(*fpaths, _raw_data=False):
    from .xlreader import load_workbook
    d = osp.dirname(fpaths[0])
    return {osp.relpath(fp, d).upper().replace('\\', '/'): _book2dict(
        load_workbook(fp, data_only=True, _raw_data=_raw_data)
    ) for fp in fpaths}


def _convert_complex(results):
    from ..functions import is_complex, str2complex
    res = {}
    it = sorted(sh.stack_nested_keys(results, depth=3))
    for k, v in it:
        if isinstance(v, str) and is_complex(v):
            v = str2complex(v)
            if v.imag:
                v = {'imag': v.imag, 'real': v.real}
            else:
                v = v.real
        sh.get_nested_dicts(res, *k, default=lambda: v)
    return res


def escape_char(m):
    return f"_x{ord(m.group(0)):04X}_"


def _path(*keys):
    p = ''
    n = len(keys)
    if n > 0:
        p = f'[{keys[0]}]'
    if n > 1:
        p += f'{keys[1]}'
    if n > 2:
        p += f'!{".".join(keys[2:])}'
    return p


def _f_value(value, maxlen=120):
    if isinstance(value, str) and not isinstance(value, sh.Token):
        value = f'"{value}"'
    value = str(value)
    return value if len(value) <= maxlen else value[:maxlen - 3] + "..."


def _format_errors_key(x):
    error_type, parent, diff = x
    key = (0,)
    try:
        rng = Ranges().push(parent[-1]).ranges[0]
        key = (1, int(rng['r1']), rng['n1'])
        return (tuple(parent[:-1]), key, error_type)
    except (InvalidRangeName, IndexError):
        return (tuple(parent), key, error_type)


def _format_errors(errors):
    for error_type, parent, diff in sorted(errors, key=_format_errors_key):
        if error_type == 'change':
            old, new = diff
            msg = f'Change {_path(*parent)}: {_f_value(old)} -> {_f_value(new)}'
            if isinstance(old, float) and isinstance(new, float):
                msg = f'{msg} (diff: {new - old})'
            yield msg
        elif error_type in ('add', 'remove'):
            msg = 'Addition' if error_type == 'add' else 'Deletion'
            for k, value in diff:
                if value != '':
                    yield f'{msg} {_path(*parent, k)} -> {_f_value(value)}'
        else:
            raise ValueError(f'Unknown error type: {error_type}')


class ExcelModel:
    compile_class = sh.DispatchPipe

    def __init__(self):
        self.dsp = sh.Dispatcher(name='ExcelModel')
        self.cells = {}
        self.books = {}
        self.basedir = None
        # AcquiOS fork 2026-05-29: cached cell values from a data_only=True
        # openpyxl pass on each loaded workbook. Used by Phase B's iterative
        # solver as a non-zero seed to escape trivial fixed points on
        # homogeneous SCCs (e.g. DCF revenue/expense feedback). Maps qualified
        # cell address (matching formulas-lib's internal format,
        # "'[file]SHEET'!CELL") to float. Empty when no workbook was loaded
        # from a file path. See specs/FORMULAS_ITERATIVE_CIRCULAR_SOLVER.md
        # §3.5c "Phase B seed fix".
        self._cached_values = {}

    def __call__(self, *args, **kwargs):
        return self.calculate(*args, **kwargs)

    def calculate(self, *args, **kwargs):
        _total_nodes = len(self.dsp.function_nodes)
        _calc_counter = [0]  # mutable for closure access
        _has_progress = _progress_callback is not None or _abort_flag is not None

        if _PROF:
            import re as _re
            _fn_times = _pcoll.defaultdict(lambda: {'t': 0.0, 'n': 0})
            _originals = {}
            _func_re = _re.compile(r"^=(?:\w+!)?([A-Z][A-Z0-9_.]+)\(", _re.IGNORECASE)
            for _fid, _fnode in self.dsp.function_nodes.items():
                _fn = _fnode['function']
                _fname = getattr(_fn, '__name__', type(_fn).__name__)
                if isinstance(_fn, CellWrapper):
                    _m = _func_re.match(_fname) if isinstance(_fname, str) else None
                    if _m:
                        _label = _m.group(1).upper()
                    elif isinstance(_fname, str) and '!' in _fname:
                        _label = 'ARITHMETIC'
                    else:
                        _label = 'CELL_OTHER'
                elif isinstance(_fn, RangesAssembler):
                    _label = 'RangesAssembler'
                elif isinstance(_fn, InvRangesAssembler):
                    _label = 'InvRangesAssembler'
                elif _fname == 'bypass':
                    _label = 'bypass'
                else:
                    _label = type(_fn).__name__
                _originals[_fid] = _fn

                class _Wrapper:
                    __slots__ = ('_fn', '_label', '_fn_times', '_counter', '_total', '_has_progress')
                    def __init__(self, fn, label, fn_times, counter, total, has_progress):
                        self._fn = fn
                        self._label = label
                        self._fn_times = fn_times
                        self._counter = counter
                        self._total = total
                        self._has_progress = has_progress
                    def __call__(self, *a, **kw):
                        _t = _ptime.perf_counter()
                        _r = self._fn(*a, **kw)
                        _d = _ptime.perf_counter() - _t
                        self._fn_times[self._label]['t'] += _d
                        self._fn_times[self._label]['n'] += 1
                        if self._has_progress:
                            self._counter[0] += 1
                            if self._counter[0] % 100 == 0:
                                _check_abort()
                                _report_progress('calculate', self._counter[0], self._total, self._label)
                        return _r
                    def __getattr__(self, name):
                        return getattr(self._fn, name)

                _fnode['function'] = _Wrapper(_fn, _label, _fn_times, _calc_counter, _total_nodes, _has_progress)

            _t0 = _ptime.perf_counter()

        elif _has_progress:
            _originals = {}
            for _fid, _fnode in self.dsp.function_nodes.items():
                _fn = _fnode['function']
                _originals[_fid] = _fn

                class _ProgressWrapper:
                    __slots__ = ('_fn', '_counter', '_total')
                    def __init__(self, fn, counter, total):
                        self._fn = fn
                        self._counter = counter
                        self._total = total
                    def __call__(self, *a, **kw):
                        _r = self._fn(*a, **kw)
                        self._counter[0] += 1
                        if self._counter[0] % 100 == 0:
                            _check_abort()
                            _report_progress('calculate', self._counter[0], self._total, '')
                        return _r
                    def __getattr__(self, name):
                        return getattr(self._fn, name)

                _fnode['function'] = _ProgressWrapper(_fn, _calc_counter, _total_nodes)

        _report_progress('calculate', 0, _total_nodes, '')
        result = self.dsp.dispatch(*args, **kwargs)

        if _PROF:
            _d = _ptime.perf_counter() - _t0
            _prof['calculate > dispatch (total)']['t'] += _d
            _prof['calculate > dispatch (total)']['n'] = _total_nodes
            _fn_nodes = self.dsp.function_nodes
            for _fid, _fn in _originals.items():
                _fn_nodes[_fid]['function'] = _fn
            for _k, _v in _fn_times.items():
                _prof[f'  calc > {_k}']['t'] += _v['t']
                _prof[f'  calc > {_k}']['n'] += _v['n']
        elif _has_progress:
            _fn_nodes = self.dsp.function_nodes
            for _fid, _fn in _originals.items():
                _fn_nodes[_fid]['function'] = _fn

        return result

    def compare(
            self, *fpaths, target=None, actual=None, solution=None,
            books=None, dirpath=None, tolerance=0, absolute_tolerance=.000001,
            dot_notation=False, formatted=True, **kwargs):
        from dictdiffer import diff
        if target is None:
            target = _convert_complex(_file2books(*fpaths))
        if actual is None:
            if solution is None:
                solution = self.dsp.dispatch()
            actual = _convert_complex(_res2books(self.write(
                books=books, dirpath=dirpath, solution=solution
            )))
        err = diff(
            target, actual, tolerance=tolerance, dot_notation=dot_notation,
            absolute_tolerance=absolute_tolerance, **kwargs
        )
        if formatted:
            err = tuple(_format_errors(err))
            if len(err) == 0:
                return 'No differences.'
            return '\n\nErrors({}):\n{}\n'.format(len(err), '\n'.join(err))

        return err

    def __getstate__(self):
        return {'dsp': self.dsp, 'cells': {}, 'books': {}}

    def _update_refs(self, nodes, refs):
        if nodes:
            dsp = self.dsp.get_sub_dsp(nodes)
            dsp.raises = ''
            sol = dsp({k: sh.EMPTY for k in dsp.data_nodes if k not in refs})
            refs.update({
                k: v for k, v in sol.items()
                if k in refs and isinstance(v, Ranges)
            })

    def add_references(self, book, context=None):
        refs, nodes = {}, set()
        it = book.defined_names
        it = it.values() if isinstance(it, dict) else it.definedName
        for n in it:
            if n.hidden or n.localSheetId is not None:
                continue  # Accepts only global references.
            ref = Ref(n.name.upper(), '=%s' % n.value, context).compile(
                context=context
            )
            nodes.update(ref.add(self.dsp, context=context))
            refs[ref.output] = None
            self.cells[ref.output] = ref
        self._update_refs(nodes, refs)
        return refs

    def loads(self, *file_names):
        for filename in file_names:
            self.load(filename)
            # AcquiOS fork 2026-05-29: parallel data_only=True pass to
            # populate _cached_values for Phase B seeding. Best-effort —
            # any failure here is non-fatal and just means Phase B falls
            # back to a zero seed for cells without cached values.
            try:
                self._load_cached_values(filename)
            except Exception as e:
                log.warning("Failed to load cached values from %s: %s", filename, e)
        return self

    def _load_cached_values(self, filename):
        # Populate self._cached_values from a data_only=True openpyxl pass.
        # Key format: "'[FILENAME]SHEET'!CELL" matching formulas-lib's
        # internal qualified-cell convention (see Ranges._push and the
        # output of cell.range.ranges[0]['name']).
        from openpyxl import load_workbook as _openpyxl_load
        if not isinstance(filename, str) or not osp.isfile(filename):
            return  # stream/buffer/non-file load — nothing to cache
        wb = _openpyxl_load(filename, data_only=True, read_only=True)
        # AcquiOS fork 2026-05-29: formulas-lib preserves the original
        # filename case in its qualified SCC keys (e.g. seen on Concord:
        # `'[recalc_sanitized_1052.xlsx]ANNUAL PROFORMA'!G132`) while
        # uppercasing the sheet name. Match that exact convention.
        fname = osp.basename(filename)
        n_added = 0
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            sheet_upper = sheet_name.upper()
            prefix = "'[%s]%s'!" % (fname, sheet_upper)
            for row in ws.iter_rows(values_only=False):
                for cell in row:
                    v = cell.value
                    if v is None:
                        continue
                    try:
                        fv = float(v)
                    except (TypeError, ValueError):
                        continue
                    if fv != fv or abs(fv) > 1e15:  # NaN or out-of-bounds
                        continue
                    key = prefix + cell.coordinate
                    self._cached_values[key] = fv
                    n_added += 1
        wb.close()
        import os as _os_cv
        if _os_cv.environ.get('FORMULAS_TIMING') == '1':
            log.warning("[FORMULAS_TIMING] cached_values_loaded: %d numeric "
                        "cells from %s", n_added, filename)

    def load(self, filename):
        if _PROF:
            _t0 = _ptime.perf_counter()
        book, context = self.add_book(filename)
        if _PROF:
            _d = _ptime.perf_counter() - _t0
            _prof['load > add_book']['t'] += _d
            _prof['load > add_book']['n'] += 1
            _t0 = _ptime.perf_counter()
        self.pushes(*book.worksheets, context=context)
        if _PROF:
            _d = _ptime.perf_counter() - _t0
            _prof['load > pushes (all sheets)']['t'] += _d
            _prof['load > pushes (all sheets)']['n'] += 1
        return self

    def from_ranges(self, *ranges):
        return self.complete(ranges)

    def pushes(self, *worksheets, context=None):
        _total_ws = len(worksheets)
        for _ws_idx, ws in enumerate(worksheets):
            _report_progress('compile_cells_sheet', _ws_idx, _total_ws, getattr(ws, 'title', '?'))
            self.push(ws, context=context)
        _report_progress('compile_cells_sheet', _total_ws, _total_ws, 'done')
        return self

    def push(self, worksheet, context):
        worksheet, context = self.add_sheet(worksheet, context)
        references = self.references
        formula_references = self.formula_references(context)
        formula_ranges = self.formula_ranges(context)
        external_links = self.external_links(context)
        ctx = {'external_links': external_links}
        ctx.update(context)
        cells = []
        _sheet = ctx.get('sheet', '?')
        if _PROF:
            _t0 = _ptime.perf_counter()
            _n_cells = 0
            _n_formulas = 0
        _cell_count = 0
        _total_est = (worksheet.max_row or 0) * (worksheet.max_column or 0)
        for row in worksheet.iter_rows():
            for c in row:
                if hasattr(c, 'value'):
                    cells.append(self.compile_cell(
                        c, ctx, references, formula_references
                    ))
                    if _PROF:
                        _n_cells += 1
                        if c.data_type == 'f':
                            _n_formulas += 1
                _cell_count += 1
                if _cell_count % 200 == 0:
                    _check_abort()
                    _report_progress('compile_cells', _cell_count, _total_est, _sheet)
        if _PROF:
            _d = _ptime.perf_counter() - _t0
            _k = f'push({_sheet}) > compile_cell'
            _prof[_k]['t'] += _d
            _prof[_k]['n'] += _n_cells
            _prof[f'push({_sheet}) > formula_cells']['n'] += _n_formulas
            _prof[f'push({_sheet}) > plain_cells']['n'] += _n_cells - _n_formulas
            _t0 = _ptime.perf_counter()
        for cell in cells:
            # noinspection PyTypeChecker
            self.add_cell(sh.await_result(cell), ctx, formula_ranges)
        if _PROF:
            _d = _ptime.perf_counter() - _t0
            _k = f'push({_sheet}) > add_cell'
            _prof[_k]['t'] += _d
            _prof[_k]['n'] += len(cells)
        return self

    def add_book(self, book=None, context=None, data_only=False):
        ctx = (context or {}).copy()
        are_in, get_in = sh.are_in_nested_dicts, sh.get_nested_dicts
        if isinstance(book, str):
            book = _encode_path(book).split('/')
            ctx['directory'], ctx['filename'] = '/'.join(book[:-1]), book[-1]
        if self.basedir is None:
            directory = _encode_path(ctx.get('directory') or '.')
            self.basedir = osp.abspath(_decode_path(directory))
            ctx['directory'] = ''
        if ctx['directory']:
            ctx['directory'] = _encode_path(osp.relpath(
                osp.join(self.basedir, _decode_path(ctx['directory'])),
                self.basedir
            ))
        if ctx['directory'] == '.':
            ctx['directory'] = ''
        fpath = osp.join(_decode_path(ctx['directory']), ctx['filename'])
        ctx['excel'] = _encode_path(fpath).upper()
        data = get_in(self.books, ctx['excel'])
        book = data.get(BOOK)
        if not book:
            from .xlreader import load_workbook
            if _PROF:
                _t0 = _ptime.perf_counter()
            data[BOOK] = book = load_workbook(
                osp.join(self.basedir, fpath), data_only=data_only
            )
            if _PROF:
                _d = _ptime.perf_counter() - _t0
                _prof['add_book > xlreader.load_workbook']['t'] += _d
                _prof['add_book > xlreader.load_workbook']['n'] += 1

        if 'external_links' not in data:
            fdir = osp.join(self.basedir, _decode_path(ctx['directory']))
            data['external_links'] = {
                str(i + 1): osp.split(osp.relpath(osp.realpath(osp.join(
                    fdir, _decode_path(el.file_link.Target)
                )), self.basedir))
                for i, el in enumerate(book._external_links)
                if el.file_link.Target.endswith('.xlsx')
            }
            data['external_links'] = {
                k: (_encode_path(d), f)
                for k, (d, f) in data['external_links'].items()
            }

        if 'references' not in data:
            if _PROF:
                _t0 = _ptime.perf_counter()
            context = {'external_links': data['external_links']}
            context.update(ctx)
            data['references'] = self.add_references(book, context=context)
            if _PROF:
                _d = _ptime.perf_counter() - _t0
                _prof['add_book > add_references']['t'] += _d
                _prof['add_book > add_references']['n'] += 1

        return book, ctx

    def add_sheet(self, worksheet, context):
        get_in = sh.get_nested_dicts
        if isinstance(worksheet, str):
            book = get_in(self.books, context['excel'], BOOK)
            worksheet = book[_get_name(worksheet, book.sheetnames)]

        ctx = {'sheet': worksheet.title.upper()}
        ctx.update(context)

        d = get_in(self.books, ctx['excel'], SHEETS, ctx['sheet'])
        if 'formula_references' not in d:
            try:
                formula_references = {
                    k: v['ref'] for k, v in worksheet.formula_attributes.items()
                    if v.get('t') == 'array' and 'ref' in v
                }
            except AttributeError:  # openpyxl>=3.1
                formula_references = worksheet.array_formulae.copy()

            d['formula_references'] = formula_references
        else:
            formula_references = d['formula_references']

        if 'formula_ranges' not in d:
            d['formula_ranges'] = {
                Ranges().push(ref, context=ctx)
                for ref in formula_references.values()
            }
        return worksheet, ctx

    @property
    def references(self):
        return sh.combine_dicts(*(
            d.get('references', {}) for d in self.books.values()
        ))

    def formula_references(self, ctx):
        return sh.get_nested_dicts(
            self.books, ctx['excel'], SHEETS, ctx['sheet'], 'formula_references'
        )

    def formula_ranges(self, ctx):
        return sh.get_nested_dicts(
            self.books, ctx['excel'], SHEETS, ctx['sheet'], 'formula_ranges'
        )

    def external_links(self, ctx):
        return sh.get_nested_dicts(self.books, ctx['excel'], 'external_links')

    @staticmethod
    def _compile_cell(crd, val, context, check_formula, references):
        cell = Cell(crd, val, context=context, check_formula=check_formula)
        cell.compile(references=references, context=context)
        return cell

    def compile_cell(self, cell, context, references, formula_references):
        crd = cell.coordinate
        crd = formula_references.get(crd, crd)
        val = cell.value

        if cell.data_type == 'f' and _is_deactivated_cell(cell):
            frozen = _frozen_value_for(cell)
            if _PROF:
                _prof['deactivated_cells']['n'] += 1
            return self._compile_cell(
                crd, frozen, context, check_formula=False,
                references=references
            )

        if cell.data_type == 'f':
            if not isinstance(val, str):
                # Excel Data Table (What-If) cells are stored as DataTableFormula
                # objects with no `.text`; the `=TABLE()` construct is unsupported.
                # They are terminal sensitivity-grid display cells (nothing reads
                # from them), so substitute an empty constant rather than crash the
                # whole workbook load. Without this, one such cell aborts the model.
                if not hasattr(val, 'text'):
                    log.warning(
                        'Data Table cell `%s` unsupported by formulas '
                        '(=TABLE() construct); substituted empty value.', crd
                    )
                    return self._compile_cell(
                        crd, None, context, check_formula=False,
                        references=references
                    )
                val = val.text
            val = val[:2] == '==' and val[1:] or val
        elif cell.data_type == 'n' and isinstance(val, float):
            val = round(val, 15)

        check_formula = cell.data_type != 's'
        return self._compile_cell(crd, val, context, check_formula, references)

    def add_cell(self, cell, context, formula_ranges):
        if cell.output in self.cells:
            return
        if cell.value is not sh.EMPTY:
            if any(not (cell.range - rng).ranges for rng in formula_ranges):
                return

        if cell.add(self.dsp, context=context):
            self.cells[cell.output] = cell
            return cell

    def add_anchor(self, rng, data_nodes=None, context=None, set_ref=True):
        if rng.get('anchor'):
            n_id = rng['name']
            name = n_id[:-1] + ':'
            ref = None
            if context:
                ref = self.formula_references(context).get(n_id)

            if ref is None and data_nodes:
                for k in data_nodes:
                    if k.startswith(name):
                        for fn in self.dsp.dmap.pred.get(k):
                            if not isinstance(
                                    self.dsp.nodes[fn]['function'],
                                    RangesAssembler
                            ):
                                ref = k
                                break
                        break
                if ref is None:
                    k = name[:-1]
                    if data_nodes.get(k):
                        for fn in self.dsp.dmap.pred.get(k):
                            if not isinstance(
                                    self.dsp.nodes[fn]['function'],
                                    RangesAssembler
                            ):
                                ref = k
                                break
            if ref:
                if ref in self.cells:
                    ref = self.cells[ref].range.ranges[0]['name']
                else:
                    ref = Ranges.get_range(ref)['name']
                self.dsp.add_function(
                    function_id=f'={ref}',
                    function=sh.bypass,
                    inputs=[ref],
                    outputs=[n_id]
                )
            elif set_ref:
                Cell(n_id, '=#REF!').compile().add(self.dsp)
            return True
        return False

    def anchors(self, stack=None):
        done = set(self.cells)
        data_nodes = self.dsp.data_nodes
        if stack is None:
            pred = self.dsp.dmap.pred
            stack = {
                k for k in data_nodes if not pred.get(k) and k.endswith('#')
            }
            stack = stack.difference(done)

        for n_id in sorted(stack):
            rng = Ranges.get_range(n_id, raise_anchor=False)
            self.add_anchor(rng, data_nodes)

    def complete(self, stack=None):
        done = set(self.cells)
        if stack is None:
            stack = {k for k in self.dsp.data_nodes if k not in self.references}
            stack = stack.difference(done)
        stack = sorted(stack)
        sheet_limits = {}
        _complete_iter = 0
        with tqdm.tqdm(total=len(stack)) as pbar:
            while stack:
                n_id = stack.pop()
                pbar.update(1)
                _complete_iter += 1
                if _complete_iter % 10 == 0:
                    _check_abort()
                    _report_progress('build_graph', pbar.n, pbar.total, '')
                if isinstance(n_id, sh.Token) or n_id in done:
                    continue
                done.add(n_id)
                if n_id in self.references:
                    extra = self.cells[n_id].inputs or ()
                    if extra:
                        stack.extend(extra)
                        pbar.total += len(extra)
                        pbar.refresh()
                    continue
                try:
                    rng = Ranges.get_range(n_id, raise_anchor=False)
                except InvalidRangeName:  # Missing Reference.
                    log.warning('Missing Reference `{}`!'.format(n_id))
                    Ref(n_id, '=#REF!').compile().add(self.dsp)
                    continue
                book = _encode_path(osp.join(
                    _decode_path(rng.get('directory', '')),
                    _decode_path(rng.get('filename', rng.get('excel_id', '')))
                ))

                try:
                    context = self.add_book(book)[1]
                    wk, context = self.add_sheet(rng['sheet'], context)
                except Exception as ex:  # Missing excel file or sheet.
                    log.warning('Error in loading `{}`:\n{}'.format(n_id, ex))
                    Cell(n_id, '=#REF!').compile().add(self.dsp)
                    self.books.pop(book, None)
                    continue
                if self.add_anchor(rng, set_ref=False, context=context):
                    continue
                references = self.references
                formula_references = self.formula_references(context)
                formula_ranges = self.formula_ranges(context)
                external_links = self.external_links(context)

                _name = '%s'
                if 'sheet_id' in rng:
                    _name = f'{rng["sheet_id"]}!{_name}'
                if wk not in sheet_limits:
                    sheet_limits[wk] = wk.max_row, wk.max_column
                max_row, max_column = sheet_limits[wk]
                it = wk.iter_rows(
                    int(rng['r1']), min(int(rng['r2']), max_row),
                    rng['n1'], min(rng['n2'], max_column)
                )
                ctx = {'external_links': external_links}
                ctx.update(context)
                cells = []
                _inner_count = 0
                for row in it:
                    for c in row:
                        _inner_count += 1
                        if _inner_count % 100 == 0:
                            _check_abort()
                            _report_progress('build_graph_cells', _inner_count, 0, str(pbar.n))
                        n = _name % c.coordinate
                        if n in self.cells:
                            continue
                        elif hasattr(c, 'value'):
                            cells.append(self.compile_cell(
                                c, ctx, references, formula_references
                            ))
                _add_count = 0
                for cell in cells:
                    _add_count += 1
                    if _add_count % 50 == 0:
                        _check_abort()
                        _report_progress('build_graph_add', _add_count, len(cells), str(pbar.n))
                    # noinspection PyTypeChecker
                    cell = self.add_cell(sh.await_result(cell), ctx,
                                         formula_ranges)
                    if cell:
                        extra = cell.inputs or ()
                        if extra:
                            stack.extend(extra)
                            pbar.total += len(extra)
                            pbar.refresh()
        return self

    def _assemble_ranges(self, cells, nodes=None, compact=1):
        get, dsp = sh.get_nested_dicts, self.dsp
        pred = dsp.dmap.pred
        if nodes is None:
            nodes = {k for k in dsp.data_nodes if k not in dsp.default_values}
        it = (
            k for k in nodes
            if not pred[k] and not isinstance(k, sh.Token)
        )
        ranges = []
        for n_id in it:
            try:
                ra = RangesAssembler(n_id, compact=compact)
            except ValueError:
                continue
            rng = ra.range.ranges[0]
            for out, idx in get(cells, 'range', rng['sheet_id'], default=list):
                if not ra.push(idx, out):
                    break
            ra.push(get(cells, 'cell', rng['sheet_id']))
            ranges.append(ra)
        ranges = sorted(ranges, key=lambda x: len(x.missing))
        for ra in ranges:
            ra.add(dsp)

    def assemble(self, compact=1):
        cells, get = {}, sh.get_nested_dicts
        for c in self.cells.values():
            if isinstance(c, Ref):
                continue
            rng = c.range.ranges[0]
            if rng.get('anchor'):
                continue
            indices = RangesAssembler._range_indices(c.range)
            if len(indices) == 1:
                get(cells, 'cell', rng['sheet_id'])[list(indices)[0]] = c.output
            else:
                get(cells, 'range', rng['sheet_id'], default=list).append(
                    (c.output, indices)
                )

        self._assemble_ranges(cells, compact=compact)
        return self

    def inverse_references(self):
        dsp = self.dsp
        pred, succ, nodes = dsp.dmap.pred, dsp.dmap.succ, dsp.nodes
        for c in tuple(self.cells.values()):
            if isinstance(c, Ref) and c.inputs:
                if c.func.dsp.function_nodes:
                    continue
                inp = c.output
                if set(pred[inp]) == {c.func.function_id}:
                    out = list(c.inputs)[0]
                    if not any(out in succ[k] for k in succ[inp]):
                        dsp.add_function(
                            '=%s' % inp, sh.bypass, inputs=[inp], outputs=[out]
                        )
                        d = nodes[inp]
                        d['inv-data'] = {out}
                        if 'filters' in nodes[out]:
                            sh.get_nested_dicts(
                                d, 'filters', default=list
                            ).extend(nodes[out]['filters'])

    def finish(self, complete=True, circular=False, assemble=True,
               anchors=True, iterate=False, iter_max=None, iter_tol=None):
        if complete:
            if _PROF:
                _t0 = _ptime.perf_counter()
            self.complete()
            if _PROF:
                _d = _ptime.perf_counter() - _t0
                _prof['finish > complete']['t'] += _d
                _prof['finish > complete']['n'] += 1
        if anchors:
            if _PROF:
                _t0 = _ptime.perf_counter()
            self.anchors()
            if _PROF:
                _d = _ptime.perf_counter() - _t0
                _prof['finish > anchors']['t'] += _d
                _prof['finish > anchors']['n'] += 1
        if assemble:
            if _PROF:
                _t0 = _ptime.perf_counter()
            self.assemble()
            if _PROF:
                _d = _ptime.perf_counter() - _t0
                _prof['finish > assemble']['t'] += _d
                _prof['finish > assemble']['n'] += 1
        if circular:
            self.solve_circular(iterate=iterate, iter_max=iter_max,
                                iter_tol=iter_tol)
        if _PROF:
            _t0 = _ptime.perf_counter()
        self.inverse_references()
        if _PROF:
            _d = _ptime.perf_counter() - _t0
            _prof['finish > inverse_references']['t'] += _d
            _prof['finish > inverse_references']['n'] += 1
        return self

    def to_dict(self):
        nodes = {
            k: d['value']
            for k, d in self.dsp.default_values.items()
            if not isinstance(k, sh.Token)
        }
        nodes = {
            k: isinstance(v, str) and v.startswith('=') and '="%s"' % v or v
            for k, v in nodes.items()
        }
        nodes = {
            k: '#EMPTY' if v == [[sh.EMPTY]] else v
            for k, v in nodes.items()
        }
        nodes = {
            k: v
            for k, v in nodes.items()
        }
        for d in self.dsp.function_nodes.values():
            fun = d['function']
            if isinstance(fun, CellWrapper):
                nodes.update(dict.fromkeys(d['outputs'], fun.__name__))
        return nodes

    def from_dict(self, adict, context=None, assemble=True, ref=True):
        refs, cells, nodes, get = {}, {}, set(), sh.get_nested_dicts
        for k, v in adict.items():
            if isinstance(v, str) and v.upper() == '#EMPTY':
                v = [[sh.EMPTY]]
            try:
                cell = Cell(k, v, context=context, replace_missing_ref=ref)
            except ValueError:
                cell = Ref(k, v, context=context).compile(context=context)
                refs[cell.output] = None
                nodes.update(cell.add(self.dsp, context=context))
            cells[cell.output] = cell
        self._update_refs(nodes, refs)
        for k, cell in cells.items():
            if k not in refs:
                nodes.update(cell.compile(references=refs).add(self.dsp))
        self.cells.update(cells)
        if assemble:
            self.assemble()
        self.inverse_references()
        return self

    def write(self, books=None, solution=None, dirpath=None):
        books = {} if books is None else books
        solution = self.dsp.solution if solution is None else solution
        are_in, get_in = sh.are_in_nested_dicts, sh.get_nested_dicts
        for k, r in solution.items():
            if isinstance(k, sh.Token):
                continue
            if isinstance(r, Ranges):
                rng = {k: v for k, v in _re_sheet_id.match(
                    r.ranges[0]['sheet_id']
                ).groupdict().items() if v is not None}
                rng.update(r.ranges[0])
            else:
                try:
                    r = Ranges().push(k, r)
                    rng = r.ranges[0]
                except ValueError:  # Reference.
                    rng = {'sheet': ''}
            fpath = _encode_path(osp.join(
                _decode_path(rng.get('directory', '')), rng.get('filename', '')
            ))
            fpath, sheet_name = _get_name(fpath, books), rng.get('sheet')
            if not (fpath and sheet_name):
                log.info('Node `%s` cannot be saved '
                         '(missing filename and/or sheet_name).' % k)
                continue
            elif _re_build_id.match(fpath):
                continue
            if not are_in(books, fpath, BOOK):
                from openpyxl import Workbook
                book = get_in(books, fpath, BOOK, default=Workbook)
                for ws in book.worksheets:
                    book.remove(ws)
            else:
                book = books[fpath][BOOK]

            sheet_names = book.sheetnames
            sheet_name = _get_name(sheet_name, sheet_names)
            if sheet_name not in sheet_names:
                book.create_sheet(sheet_name)
            sheet = book[sheet_name]
            rng['c1'] = rng['c1'] or 'A'
            rng['r1'] = int(rng['r1']) or 1
            ref = '{c1}{r1}:{c2}{r2}'.format(**rng)
            for c, v in zip(np.ravel(sheet[ref]), np.ravel(r.value)):
                try:
                    if isinstance(v, Ranges) and v.value.shape == (1, 1):
                        v = v.value[0, 0]
                    if v is sh.EMPTY:
                        v = None
                    elif isinstance(v, np.generic):
                        v = v.item()
                    elif isinstance(v, XlError):
                        v = str(v)
                    if isinstance(v, str):
                        v = v.replace('\r', '_x000D_')
                    try:
                        c.value = v
                    except IllegalCharacterError:
                        c.value = ILLEGAL_CHARACTERS_RE.sub(escape_char, v)
                    if c.data_type == 'f':
                        c.data_type = 's'
                except AttributeError:
                    pass
        if dirpath:
            os.makedirs(dirpath, exist_ok=True)
            for fpath, d in books.items():
                d[BOOK].save(osp.join(dirpath, _decode_path(fpath)))
        return books

    def compile(self, inputs, outputs):
        dsp = self.dsp.shrink_dsp(inputs=inputs, outputs=outputs)
        inp = set(inputs)
        nodes = dsp.nodes
        for i in inputs:
            inp.update(nodes.get(i, {}).get('inv-data', ()))
        dsp.default_values = {
            k: v for k, v in dsp.default_values.items() if k not in inp
        }

        res = dsp()

        dsp = dsp.get_sub_dsp_from_workflow(
            outputs, graph=dsp.dmap, reverse=True, blockers=res,
            wildcard=False
        )

        for k, v in res.items():
            if k in dsp.data_nodes and k not in dsp.default_values:
                dsp.set_default_value(k, v.value)

        func = self.compile_class(
            dsp=dsp,
            function_id=self.dsp.name,
            inputs=inputs,
            outputs=outputs
        )

        return func

    def solve_circular(self, iterate=False, iter_max=None, iter_tol=None):
        # Phase A — graph-rewriting cycle handler.
        # ORIGINAL behavior: enumerate every elementary cycle in dmap.succ
        # (Johnson's algorithm via simple_cycles), then iterate per-cycle and
        # try to "cut" each one by rewriting an IF whose non-cycle branch is
        # statically safe. Cells where rewrite fails get ERR_CIRCULAR.
        #
        # AcquiOS fork 2026-05-29: Johnson's is O((V+E)(C+1)) where C is the
        # elementary-cycle count. On a 192K-formula DCF model with one 800-
        # cell SCC (Concord Remington, 110 back-edges) C is exponential and
        # simple_cycles ran >30 min without returning (FORMULAS_TIMING=1 H1.1
        # data; specs/FORMULAS_ITERATIVE_CIRCULAR_SOLVER.md §3.5c).
        #
        # Fix: SCC-first dispatch. Discover SCCs in O(V+E) using the iterative
        # Tarjan we already have. For each SCC of size > 1, run simple_cycles
        # only WITHIN that SCC subgraph, with a hard cycle-count cap. If an
        # SCC contains too many cycles to enumerate safely, skip the rewrite
        # pass entirely and mark all its data nodes as unbreakable — Phase B
        # iterative solver then handles it. This preserves the per-elementary-
        # cycle rewrite semantics on small SCCs (where the original test
        # fixtures live) and degrades gracefully on large ones.
        from .cycle import simple_cycles, _strongly_connected_components
        from collections import Counter
        import os as _os, time as _time
        _timing = _os.environ.get('FORMULAS_TIMING') == '1'
        # Cap on elementary cycles enumerated per SCC. Above this, fall back
        # to "whole SCC unbreakable" and let Phase B iterate. Env-overridable.
        _cycle_cap = int(_os.environ.get('FORMULAS_PHASE_A_CYCLE_CAP', '10000'))
        def _t(label, t0):
            if _timing:
                log.warning("[FORMULAS_TIMING] %s: %.3fs", label, _time.perf_counter() - t0)
        _t0_total = _time.perf_counter()
        mod, dsp = {}, self.dsp
        f_nodes, d_nodes, dmap = dsp.function_nodes, dsp.data_nodes, dsp.dmap
        skip_nodes = {
            k for k, node in f_nodes.items()
            if isinstance(node['function'], InvRangesAssembler)
        }
        _t('PhaseA.skip_nodes_build', _t0_total)

        # SCC discovery over the full dependency graph minus skip_nodes.
        _t0 = _time.perf_counter()
        full_succ = {
            v: set(nbrs) - skip_nodes
            for v, nbrs in dmap.succ.items()
            if v not in skip_nodes
        }
        sccs = [s for s in _strongly_connected_components(full_succ) if len(s) > 1]
        _t('PhaseA.scc_discovery', _t0)
        if _timing:
            log.warning("[FORMULAS_TIMING] PhaseA.scc_discovery: %d SCCs of size > 1", len(sccs))

        # Per-SCC cycle enumeration + IF-rewrite.
        _t0 = _time.perf_counter()
        unbreakable_cells = set()
        cycles_total = 0
        sccs_capped = 0
        for scc_list in sccs:
            scc_set = set(scc_list)
            # simple_cycles within just this SCC's subgraph (bounded by SCC size).
            scc_succ = {v: full_succ[v] & scc_set for v in scc_set}
            # Enumerate cycles up to the cap; if we hit the cap, bail on this
            # SCC's rewrite and mark all data nodes unbreakable.
            cycles = []
            capped = False
            for cyc in simple_cycles(scc_succ, copy=False):
                cycles.append(cyc)
                if len(cycles) > _cycle_cap:
                    capped = True
                    break
            if capped:
                sccs_capped += 1
                if _timing:
                    log.warning(
                        "[FORMULAS_TIMING] PhaseA.scc_capped: SCC size %d exceeded %d-cycle cap, "
                        "marking %d data nodes unbreakable",
                        len(scc_set), _cycle_cap, len(scc_set & set(d_nodes))
                    )
                dist = sh.inf(len(scc_set) + 1, 0)
                for k in sorted(scc_set.intersection(d_nodes)):
                    dsp.set_default_value(k, ERR_CIRCULAR, dist)
                    unbreakable_cells.add(k)
                continue
            cycles_total += len(cycles)
            # AcquiOS fork 2026-05-29: when iterate=True, skip the IF-rewrite
            # entirely. The rewrite substitutes ERR_CIRCULAR for cyclic IF
            # branches; XlCircular.__str__ returns '0', so arithmetic predecessors
            # silently coerce the error to 0, and IF(true, 0, safe_branch) returns
            # 0 — masking what should have been an iterative cycle. With iterate=True
            # the user has opted into Phase B, so leave the cycle intact and let
            # Phase B compute the real fixed point. With iterate=False, preserve
            # the legacy parameterized-function behavior (test_excel_model_compile
            # asserts func(True) returns ERR_CIRCULAR).
            if iterate:
                dist = sh.inf(len(scc_set) + 1, 0)
                for k in sorted(scc_set.intersection(d_nodes)):
                    dsp.set_default_value(k, ERR_CIRCULAR, dist)
                    unbreakable_cells.add(k)
                continue
            # Per-elementary-cycle rewrite (preserves original semantics).
            cycles_nodes = Counter(sum(cycles, []))
            for cycle in sorted(map(set, cycles)):
                cycles_nodes.subtract(cycle)
                active_nodes = {k for k, v in cycles_nodes.items() if v}
                for k in sorted(cycle.intersection(f_nodes)):
                    if _check_cycles(dmap, k, f_nodes, cycle, active_nodes, mod):
                        break
                else:
                    cycles_nodes.update(cycle)
                    dist = sh.inf(len(cycle) + 1, 0)
                    for k in sorted(cycle.intersection(d_nodes)):
                        dsp.set_default_value(k, ERR_CIRCULAR, dist)
                        unbreakable_cells.add(k)
        _t('PhaseA.if_rewrite_pass', _t0)
        if _timing:
            log.warning(
                "[FORMULAS_TIMING] PhaseA.cycles_enumerated_total: %d (across %d SCCs, %d capped)",
                cycles_total, len(sccs), sccs_capped
            )

        _t0 = _time.perf_counter()
        if mod:  # Update dsp.
            dsp.add_data(CIRCULAR, ERR_CIRCULAR)

            for k, v in mod.items():
                d = f_nodes[k]
                d['inputs'] = [CIRCULAR if i in v else i for i in d['inputs']]
                dmap.remove_edges_from(((i, k) for i in v))
                dmap.add_edge(CIRCULAR, k)
        _t('PhaseA.dsp_mutation', _t0)
        _t('PhaseA.TOTAL', _t0_total)

        # Phase B — opt-in iterative solver.
        # AcquiOS fork, 2026-05-29. See specs/FORMULAS_ITERATIVE_CIRCULAR_SOLVER.md.
        # For each SCC of "unbreakable" cells (where Phase A's IF-rewrite
        # couldn't statically dissolve the cycle), iterate to a fixed point
        # using scipy.optimize.fixed_point, with broyden1 and plain Gauss-
        # Seidel as fallbacks. Writes converged values back to dsp.default_values
        # so they appear in subsequent .calculate() output.
        self._last_solve_info = {'sccs': [], 'iterate_enabled': iterate}
        if not iterate or not unbreakable_cells:
            return self

        # Build dependency map restricted to the unbreakable cells (and their
        # function nodes' predecessors), then take SCCs of size > 1.
        sub_succ = {
            v: set(nbrs) - skip_nodes
            for v, nbrs in dmap.succ.items()
            if v not in skip_nodes
        }
        sccs = [
            s for s in _strongly_connected_components(sub_succ)
            if len(s) > 1 and any(c in unbreakable_cells for c in s)
        ]
        if not sccs:
            return self

        # Read iterative-calc config from the workbook with Excel defaults.
        wb_iter_count, wb_iter_delta = self._read_workbook_iterate_config()
        max_iter = iter_max or wb_iter_count or 100
        tolerance = iter_tol or wb_iter_delta or 0.001

        for raw_scc in sccs:
            # Restrict to data nodes (cell addresses) that have a corresponding
            # Cell object with .func. Function nodes inside the SCC are reached
            # via cell.func, not solved directly.
            scc = [c for c in raw_scc if c in self.cells and self.cells[c].func]
            if not scc:
                continue
            self._solve_scc_iteratively(scc, max_iter, tolerance)

        return self

    def _read_workbook_iterate_config(self):
        # Returns (iterateCount, iterateDelta) from the first workbook's
        # calcPr, or (None, None) if unavailable.
        try:
            wb = next(iter(self.books.values())).get(BOOK)
            calc = wb.calculation
            return calc.iterateCount, calc.iterateDelta
        except (AttributeError, StopIteration, KeyError):
            return None, None

    def _solve_scc_iteratively(self, scc, max_iter, tolerance):
        # scc is a list of qualified cell addresses (e.g. "'[wb]Sheet'!A2").
        # Each cell has a CellWrapper at self.cells[c].func that takes one
        # arg per entry in self.cells[c].inputs. Args may be either scalar
        # Ranges (single-cell inputs) OR multi-cell Ranges (range inputs to
        # INDEX/MATCH/VLOOKUP/SUM-style functions).
        import numpy as np
        from scipy.optimize import fixed_point, broyden1

        scc_set = set(scc)
        scc_idx = {c: i for i, c in enumerate(scc)}

        # AcquiOS fork 2026-05-29: pre-compute the full workbook solution once
        # so we have correctly-assembled Ranges objects for ALL non-SCC inputs
        # the SCC depends on — including ranges like Refs!A1:A3 that need
        # RangesAssembler to build. The prior approach read scalar default_values
        # which silently returned 0 for range keys, making INDEX/MATCH inside
        # cyclic IF branches evaluate to 0 (Test 4 toggle=1 reproduces).
        try:
            pre_sol = self.calculate()
        except Exception as e:
            log.warning("Phase B pre-solve calculate() failed: %s — falling back "
                        "to scalar default reads", e)
            pre_sol = {}

        # Build frozen inputs from pre_sol when possible, falling back to scalar
        # default reads for keys not in the solution.
        frozen = {}
        for c in scc:
            cell_obj = self.cells[c]
            for k in cell_obj.inputs:
                if k in scc_set or k in frozen:
                    continue
                if k in pre_sol:
                    frozen[k] = pre_sol[k]  # already a Ranges object
                else:
                    frozen[k] = _make_scalar_ranges(
                        k, _read_scalar_from_dsp(self.dsp, k)
                    )

        def g(x):
            # Snapshot for this iteration: frozen inputs as-is, SCC cells
            # wrapped as scalar Ranges with the current iterate.
            snap = dict(frozen)
            for c, v in zip(scc, x):
                snap[c] = _make_scalar_ranges(c, float(v))
            out = np.zeros(len(scc))
            for c in scc:
                cell_obj = self.cells[c]
                in_keys = list(cell_obj.inputs)
                args = [snap[k] for k in in_keys]
                try:
                    res = cell_obj.func(*args)
                except Exception:
                    # Cell function blew up on this iterate — leave value as-is.
                    out[scc_idx[c]] = _to_scalar(snap[c])
                    continue
                out[scc_idx[c]] = _to_scalar(res)
            return out

        # AcquiOS fork 2026-05-29: seed from cached values when available;
        # fall back to 1.0 (NOT 0.0) for cells without cached values. Many
        # DCF SCCs are homogeneous (g(0)=0 is a valid trivial fixed point)
        # so the zero seed traps Steffensen on the wrong solution. A
        # non-zero seed breaks the homogeneous symmetry. Final fallback is
        # zero only if 1.0 produces non-finite g() output.
        x0 = np.ones(len(scc))
        n_seeded = 0
        for i, c in enumerate(scc):
            v = self._cached_values.get(c)
            if v is not None:
                x0[i] = v
                n_seeded += 1
        import os as _os_seed
        if _os_seed.environ.get('FORMULAS_TIMING') == '1':
            log.warning("[FORMULAS_TIMING] Phase_B_seed: %d/%d cells seeded "
                        "from cached values, rest = 1.0", n_seeded, len(scc))
        converged, x_star, method_used = False, x0, None

        # AcquiOS fork 2026-05-29 — diagnostic instrumentation (env-gated).
        # FORMULAS_SCC_TRACE=A1,A2,... dumps the SCC trajectory for cells whose
        # qualified address contains any of those substrings. Logs: x0 value,
        # final x_star, g(x_star), residual. Helps diagnose "converged but to
        # wrong fixed point" failure modes (e.g. trivial zero fixed point).
        import os as _os_diag
        _trace_keys = [k.strip() for k in
                       _os_diag.environ.get('FORMULAS_SCC_TRACE', '').split(',')
                       if k.strip()]
        if _trace_keys:
            _trace_idx = [
                (i, c) for i, c in enumerate(scc)
                if any(tk.upper() in c.upper() for tk in _trace_keys)
            ]
            if _trace_idx:
                log.warning("[FORMULAS_SCC_TRACE] SCC size=%d, tracing %d cells",
                            len(scc), len(_trace_idx))
                gx0 = g(x0)
                for i, c in _trace_idx:
                    log.warning("[FORMULAS_SCC_TRACE]   x0[%s] = %g, g(x0)[%s] = %g",
                                c, x0[i], c, gx0[i])
        else:
            _trace_idx = None

        # Sanity bound: DCF values rarely exceed 1e12 (a trillion). Anything
        # beyond this is broyden1 / fixed_point marching off to infinity on a
        # divergent system — a relative-residual check can't catch this because
        # at huge |x| even non-zero residuals look proportionally tiny.
        DIVERGENCE_BOUND = 1e15

        def _verify(x):
            # Post-hoc convergence check. Must reject:
            #  - non-finite values
            #  - values that exploded past DIVERGENCE_BOUND (divergent system)
            #  - residuals (in absolute terms) larger than tolerance
            if not np.all(np.isfinite(x)):
                return False
            if np.max(np.abs(x)) > DIVERGENCE_BOUND:
                return False
            gx = g(x)
            if not np.all(np.isfinite(gx)):
                return False
            abs_res = np.max(np.abs(gx - x))
            rel_res = np.max(np.abs(gx - x) / np.maximum(np.abs(x), 1.0))
            # Accept if either absolute or relative residual is tight.
            return abs_res < tolerance * 100 or rel_res < tolerance

        # Tier 1: scipy.optimize.fixed_point (Steffensen + Aitken Δ²).
        try:
            cand = fixed_point(g, x0, xtol=tolerance, maxiter=max_iter,
                               method='del2')
            if _verify(cand):
                x_star, converged, method_used = cand, True, 'fixed_point'
        except RuntimeError:
            pass

        # Tier 2: scipy.optimize.broyden1 (quasi-Newton).
        if not converged:
            try:
                with np.errstate(divide='ignore', invalid='ignore'):
                    cand = broyden1(lambda x: g(x) - x, x0,
                                    f_tol=tolerance, maxiter=max_iter)
                if _verify(cand):
                    x_star, converged, method_used = cand, True, 'broyden'
            except Exception:
                pass

        # Tier 3: plain Gauss-Seidel fallback.
        if not converged:
            prev = x0.copy()
            for i in range(max_iter):
                new = g(prev)
                if not np.all(np.isfinite(new)):
                    # Divergence — bail and report.
                    x_star, method_used = prev, 'gauss_seidel_diverged'
                    break
                rel_change = np.max(
                    np.abs(new - prev) / np.maximum(np.abs(prev), 1.0)
                )
                if rel_change < tolerance:
                    converged, method_used = True, 'gauss_seidel'
                    x_star = new
                    break
                prev = new
            else:
                # Max iterations hit without convergence — return last iterate.
                x_star, method_used = prev, 'gauss_seidel_unconverged'

        # Post-solve trace (FORMULAS_SCC_TRACE).
        if _trace_idx:
            gxstar = g(x_star)
            log.warning("[FORMULAS_SCC_TRACE] post-solve: method=%s, converged=%s",
                        method_used, converged)
            for i, c in _trace_idx:
                resid = gxstar[i] - x_star[i]
                log.warning(
                    "[FORMULAS_SCC_TRACE]   x_star[%s] = %g, g(x_star)[%s] = %g, "
                    "residual = %g", c, x_star[i], c, gxstar[i], resid
                )

        # Write converged values back to the dispatcher so they appear in
        # subsequent .calculate() output. Override the ERR_CIRCULAR default
        # value set by Phase A.
        for c, v in zip(scc, x_star):
            try:
                self.dsp.set_default_value(c, _make_scalar_ranges(c, float(v)))
            except Exception as e:
                log.warning("Failed to write converged value for %s: %s", c, e)

        self._last_solve_info['sccs'].append({
            'cells': list(scc),
            'size': len(scc),
            'converged': converged,
            'method': method_used,
            'final_values': {c: float(v) for c, v in zip(scc, x_star)},
        })


def _check_range_all_cycles(nodes, active_nodes, j):
    if isinstance(nodes[j]['function'], RangesAssembler):
        return active_nodes.intersection(nodes[j]['inputs'])
    return False


def _check_cycles(dmap, node_id, nodes, cycle, active_nodes, mod=None):
    node, mod = nodes[node_id], {} if mod is None else mod
    _map = dict(zip(node['function'].inputs, node['inputs']))
    pred, res = dmap.pred, ()
    check = functools.partial(_check_range_all_cycles, nodes, active_nodes)
    if not any(any(map(check, pred[k])) for k in _map.values() if k in cycle):
        cycle = [i for i, j in _map.items() if j in cycle]
        try:
            res = tuple(map(_map.get, node['function'].check_cycles(cycle)))
            res and sh.get_nested_dicts(mod, node_id, default=set).update(res)
        except AttributeError:
            pass
    return res

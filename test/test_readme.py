#!/usr/bin/env python
# -*- coding: UTF-8 -*-
#
# Copyright 2016-2026 European Commission (JRC);
# Licensed under the EUPL (the 'Licence');
# You may not use this work except in compliance with the Licence.
# You may obtain a copy of the Licence at: http://ec.europa.eu/idabc/eupl
import doctest
import os
import unittest
from pathlib import Path

from openpyxl import load_workbook

EXTRAS = os.environ.get('EXTRAS', 'all')
ROOT = Path(__file__).resolve().parents[1]
README = ROOT / 'README.rst'


@unittest.skipIf(EXTRAS not in ('all',), 'Not for extra %s.' % EXTRAS)
class TestReadme(unittest.TestCase):
    def test_readme(self):
        failure_count, test_count = doctest.testfile(
            '../README.rst', optionflags=doctest.NORMALIZE_WHITESPACE | doctest.ELLIPSIS
        )
        self.assertGreater(test_count, 0, (failure_count, test_count))
        self.assertEqual(failure_count, 0, (failure_count, test_count))

    def test_excel_function_coverage_total_matches_workbook(self):
        content = README.read_text()
        wb = load_workbook(ROOT / 'test' / 'test_files' / 'test.xlsx', data_only=True, read_only=True)
        ws = wb[wb.sheetnames[0]]
        implemented = ws['B18'].value
        total = ws['D18'].value
        coverage = ws['E18'].value

        self.assertEqual(483, implemented)
        self.assertEqual(536, total)
        self.assertAlmostEqual(0.9011194029850746, coverage)

        self.assertIn('| TOTAL          | 483         | 536   | 90.1%    |', content)
        self.assertIn('Overall coverage is currently 483 out of 536 functions (90.1%).', content)

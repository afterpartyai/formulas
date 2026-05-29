"""Generate synthetic single-SCC workbooks of configurable size for solver
scaling tests (per FORMULAS_ITERATIVE_CIRCULAR_SOLVER.md §3.5c H2).

Each generated workbook contains exactly one SCC of size N where every cell
participates in a known-convergent linear cycle. Three patterns:

    --pattern ring
        A1 = 0.5 * AN + 1
        A2 = 0.5 * A1 + 1
        ...
        Pure ring; minimal fan-in/fan-out; converges to 2 for any seed.

    --pattern dense
        Ai = (sum(A1..AN) / N) * 0.9 + 1
        Full fan-in per cell; converges to 10 for any seed. Stresses Jacobian
        density in scipy tiers.

    --pattern chain_with_anchor
        Ring like `ring` but every 10th cell adds an external constant input.
        Models DCF-style "feedback loop with exogenous drivers."

Usage:
    python build_synthetic_sccs.py --pattern ring --size 100 --out /tmp/scc_ring_100.xlsx

Reusable: callable as `build(pattern, size, out_path)` from pytest fixtures.
"""
import argparse
import os.path as osp
import sys

try:
    from openpyxl import Workbook
except ImportError:
    print("ERROR: openpyxl required. Install with: pip install openpyxl", file=sys.stderr)
    sys.exit(2)


def _cell(row):
    return f"A{row}"


def build_ring(size, ws):
    """Pure ring: A_i = 0.5 * A_{i-1} + 1; A_1 references A_N."""
    for i in range(1, size + 1):
        prev = size if i == 1 else i - 1
        ws[_cell(i)] = f"=0.5*{_cell(prev)}+1"


def build_dense(size, ws):
    """Dense: every cell averages all cells. A_i = AVG(A1..AN)*0.9 + 1."""
    span = f"A1:A{size}"
    for i in range(1, size + 1):
        ws[_cell(i)] = f"=AVERAGE({span})*0.9+1"


def build_chain_with_anchor(size, ws):
    """Ring with periodic exogenous anchors. Every 10th cell adds a constant."""
    for i in range(1, size + 1):
        prev = size if i == 1 else i - 1
        if i % 10 == 0:
            ws[_cell(i)] = f"=0.5*{_cell(prev)}+1+{i * 0.1}"
        else:
            ws[_cell(i)] = f"=0.5*{_cell(prev)}+1"


PATTERNS = {
    'ring': build_ring,
    'dense': build_dense,
    'chain_with_anchor': build_chain_with_anchor,
}


def build(pattern, size, out_path):
    if pattern not in PATTERNS:
        raise ValueError(f"unknown pattern {pattern!r}; choose from {list(PATTERNS)}")
    wb = Workbook()
    ws = wb.active
    ws.title = 'SCC'
    PATTERNS[pattern](size, ws)
    wb.save(out_path)
    return out_path


def main():
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument('--pattern', choices=list(PATTERNS), required=True)
    ap.add_argument('--size', type=int, required=True)
    ap.add_argument('--out', required=True)
    args = ap.parse_args()
    path = build(args.pattern, args.size, args.out)
    print(f"wrote {path}: pattern={args.pattern}, size={args.size}")


if __name__ == '__main__':
    main()

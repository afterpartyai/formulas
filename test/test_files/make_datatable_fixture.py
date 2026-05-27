"""Regenerate the Data Table fixture for the =TABLE() compile-crash regression test.

Run:  python test/test_files/make_datatable_fixture.py
Produces:  test/test_files/datatable.xlsx

The fixture contains one Excel Data Table (What-If) cell (B2), stored by openpyxl
as a DataTableFormula. Before the fork patch, this crashed ExcelModel load with
`'DataTableFormula' object has no attribute 'text'`. A plain formula (A2) is
included so the test can assert the rest of the model still computes.
"""
import os.path as osp
import openpyxl
from openpyxl.worksheet.formula import DataTableFormula


def build(path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws['A1'] = 10
    ws['A2'] = '=A1*2'          # ordinary formula; must still solve to 20
    ws['B2'] = DataTableFormula(ref="B2", ca=False, dt2D=False, dtr=False, r1="A1")
    wb.save(path)
    return path


if __name__ == '__main__':
    out = osp.join(osp.dirname(__file__), 'datatable.xlsx')
    build(out)
    print("wrote", out)
